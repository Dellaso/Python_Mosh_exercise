# could finish in about 10 mins
import openpyxl as xl
from openpyxl.chart import BarChart, Reference


def process_workbook(filename):
    wb = xl.load_workbook(filename)
    sheet = wb['Sheet1']
    # cell = sheet['c1']
    # cell = sheet.cell(1, 3)
    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row, 3)
        corrected_value = cell.value * 0.9
        corrected_value_cell = sheet.cell(row, 4)
        corrected_value_cell.value = corrected_value
    
    values = Reference(sheet,
                       min_col=4,
                       min_row=2,
                       max_col=4,
                       max_row=sheet.max_row)
    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, 'e2')
    
    wb.save('filename')

